from pathlib import Path
from typing import List, Union

import openpyxl
import openpyxl as xl
from openpyxl.styles import numbers
from openpyxl.styles import builtins
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas

PATH_TYPE = Union[str, Path]
CELL_STYLE = List[str]

class ExcelWriter:
    def __init__(self, filename:PATH_TYPE, save_filename:PATH_TYPE=None, read_only:bool=False, keep_vba:bool=False, data_only:bool=False, keep_links:bool=True):
        self.workbook = xl.load_workbook(filename=filename, read_only=read_only, keep_vba=keep_vba, data_only=data_only, keep_links=keep_links)
        self.save_filename = save_filename if save_filename is not None else filename


    def __enter__(self) -> openpyxl.workbook.workbook.Workbook:
        return self


    def __exit__(self, exc_type, exc_value, tracebac):
        self.save()


    def save(self):
        self.workbook.save(self.save_filename)


    def get_worksheet(self, sheet_name:str) -> openpyxl.worksheet.worksheet.Worksheet:
        return self.workbook[sheet_name]


    def write_dataframe(self, df: pandas.DataFrame, ws: openpyxl.worksheet.worksheet.Worksheet, column_start: int, row_start: int, cell_styles: CELL_STYLE=None):
        df = df.reset_index(drop=True)

        for col_no, column in enumerate(df.columns, column_start):
            cell_style = cell_styles[column_start - col_no] if cell_styles is not None else None

            for row_no, value in enumerate(df[column], row_start):
                cell = ws.cell(row=row_no, column=col_no)
                cell.value = value
                if cell_style is not None:
                    if cell_style in builtins.styles.keys():
                        cell.style = cell_style
                    else:
                        cell.number_format = cell_style


    def write_image(self, filename: PATH_TYPE, ws: openpyxl.worksheet.worksheet.Worksheet, **args):
        pass

    def write_image_by_column_and_row(self, filename: PATH_TYPE, ws: openpyxl.worksheet.worksheet.Worksheet, col:int=None, row:int=None):
        img = Image(filename)
        if col and row:
            ws.add_image(img, ws.cell(col, row))
        else:
            ws.add_image(img)


    def write_image_by_cell_name(self, filename: PATH_TYPE, ws: openpyxl.worksheet.worksheet.Worksheet, cell:str=None):
        img = Image(filename)
        if cell:
            ws.add_image(img, cell)
        else:
            ws.add_image(img)
